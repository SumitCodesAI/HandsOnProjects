#!/usr/bin/env python3
"""
Healthcare Claims Schema Documentation Generator
Generates Excel and Markdown documentation for the healthcare_claims Databricks schema.
Falls back to realistic mock data if Databricks credentials are not available.
"""

import os
import sys
import json
from datetime import datetime, timezone
from pathlib import Path

# ── Try Databricks connection ────────────────────────────────────────────────
DATABRICKS_HOST      = os.environ.get("DATABRICKS_HOST", "")
DATABRICKS_TOKEN     = os.environ.get("DATABRICKS_TOKEN", "")
DATABRICKS_HTTP_PATH = os.environ.get("DATABRICKS_HTTP_PATH", "")
SCHEMA_NAME          = "healthcare_claims"
CATALOG_NAME         = os.environ.get("DATABRICKS_CATALOG", "hive_metastore")
OUTPUT_DIR           = Path("/home/runner/work/HandsOnProjects/HandsOnProjects/docs/databricks")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ── Mock schema definition ───────────────────────────────────────────────────
# Each table: { name, description, estimated_row_count, columns: [{...}] }
MOCK_SCHEMA = {
    "catalog": CATALOG_NAME,
    "schema":  SCHEMA_NAME,
    "source":  "sample/demo data (no Databricks credentials found)",
    "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
    "tables": [
        {
            "name": "claims",
            "description": (
                "Core fact table storing all submitted insurance claims. Each row "
                "represents one claim submitted by a provider on behalf of a patient."
            ),
            "estimated_row_count": 8_420_315,
            "columns": [
                {"name": "claim_id",            "data_type": "STRING",    "nullable": False, "primary_key": True,  "description": "Unique identifier for each claim (UUID)"},
                {"name": "claim_number",         "data_type": "STRING",    "nullable": False, "primary_key": False, "description": "Human-readable claim reference number (e.g. CLM-2024-000001)"},
                {"name": "patient_id",           "data_type": "STRING",    "nullable": False, "primary_key": False, "description": "FK → patients.patient_id"},
                {"name": "provider_id",          "data_type": "STRING",    "nullable": False, "primary_key": False, "description": "FK → providers.provider_id"},
                {"name": "claim_type",           "data_type": "STRING",    "nullable": False, "primary_key": False, "description": "Type of claim: MEDICAL, DENTAL, VISION, PHARMACY"},
                {"name": "claim_status",         "data_type": "STRING",    "nullable": False, "primary_key": False, "description": "Current status: SUBMITTED, PENDING, APPROVED, DENIED, PAID, APPEALED"},
                {"name": "service_date_start",   "data_type": "DATE",      "nullable": False, "primary_key": False, "description": "First date of service covered by this claim"},
                {"name": "service_date_end",     "data_type": "DATE",      "nullable": True,  "primary_key": False, "description": "Last date of service (NULL for single-day claims)"},
                {"name": "submission_date",      "data_type": "TIMESTAMP", "nullable": False, "primary_key": False, "description": "UTC timestamp when claim was submitted to the payer"},
                {"name": "adjudication_date",    "data_type": "TIMESTAMP", "nullable": True,  "primary_key": False, "description": "UTC timestamp when payer completed adjudication"},
                {"name": "billed_amount",        "data_type": "DECIMAL(12,2)", "nullable": False, "primary_key": False, "description": "Total amount billed by provider"},
                {"name": "allowed_amount",       "data_type": "DECIMAL(12,2)", "nullable": True,  "primary_key": False, "description": "Payer-allowed amount after contract rates applied"},
                {"name": "paid_amount",          "data_type": "DECIMAL(12,2)", "nullable": True,  "primary_key": False, "description": "Actual amount paid by insurer"},
                {"name": "patient_responsibility","data_type": "DECIMAL(12,2)","nullable": True, "primary_key": False, "description": "Copay + coinsurance + deductible owed by patient"},
                {"name": "denial_reason_code",   "data_type": "STRING",    "nullable": True,  "primary_key": False, "description": "CARC/RARC code when claim_status = DENIED"},
                {"name": "place_of_service",     "data_type": "STRING",    "nullable": False, "primary_key": False, "description": "CMS Place of Service code (e.g. 11=Office, 21=Inpatient)"},
                {"name": "npi",                  "data_type": "STRING",    "nullable": False, "primary_key": False, "description": "National Provider Identifier of billing provider"},
                {"name": "insurance_plan_id",    "data_type": "STRING",    "nullable": False, "primary_key": False, "description": "Identifier of insurance plan used for this claim"},
                {"name": "created_at",           "data_type": "TIMESTAMP", "nullable": False, "primary_key": False, "description": "Row insert timestamp (UTC)"},
                {"name": "updated_at",           "data_type": "TIMESTAMP", "nullable": False, "primary_key": False, "description": "Row last-update timestamp (UTC)"},
                {"name": "partition_date",       "data_type": "DATE",      "nullable": False, "primary_key": False, "description": "Partition column = service_date_start truncated to month"},
            ]
        },
        {
            "name": "patients",
            "description": (
                "Dimension table containing de-identified patient demographics and "
                "insurance enrollment information."
            ),
            "estimated_row_count": 1_245_890,
            "columns": [
                {"name": "patient_id",          "data_type": "STRING",  "nullable": False, "primary_key": True,  "description": "Unique patient identifier (UUID)"},
                {"name": "member_id",           "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "Insurance member ID issued by the payer"},
                {"name": "date_of_birth",       "data_type": "DATE",    "nullable": False, "primary_key": False, "description": "Patient date of birth"},
                {"name": "gender",              "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Biological sex: M, F, U (unknown)"},
                {"name": "gender_identity",     "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Self-reported gender identity"},
                {"name": "race_ethnicity_code", "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "CDC race/ethnicity code for analytics"},
                {"name": "zip_code",            "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "5-digit ZIP code (last 2 digits masked for PHI)"},
                {"name": "state_code",          "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "2-letter US state abbreviation"},
                {"name": "county_fips",         "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "5-digit FIPS county code"},
                {"name": "plan_type",           "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "Insurance plan type: HMO, PPO, EPO, HDHP, MEDICAID, MEDICARE"},
                {"name": "group_number",        "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Employer group number for commercial plans"},
                {"name": "enrollment_start",    "data_type": "DATE",    "nullable": False, "primary_key": False, "description": "Date patient enrolled in current plan"},
                {"name": "enrollment_end",      "data_type": "DATE",    "nullable": True,  "primary_key": False, "description": "Date enrollment ended; NULL if still active"},
                {"name": "primary_care_npi",    "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "NPI of assigned primary care physician"},
                {"name": "chronic_condition_flags", "data_type": "MAP<STRING,BOOLEAN>", "nullable": True, "primary_key": False, "description": "Map of CMS chronic condition flags (e.g. diabetes→true)"},
                {"name": "created_at",          "data_type": "TIMESTAMP","nullable": False, "primary_key": False, "description": "Row insert timestamp (UTC)"},
                {"name": "updated_at",          "data_type": "TIMESTAMP","nullable": False, "primary_key": False, "description": "Row last-update timestamp (UTC)"},
            ]
        },
        {
            "name": "providers",
            "description": (
                "Dimension table of healthcare providers including physicians, hospitals, "
                "clinics, and ancillary service providers."
            ),
            "estimated_row_count": 98_432,
            "columns": [
                {"name": "provider_id",         "data_type": "STRING",  "nullable": False, "primary_key": True,  "description": "Unique provider record identifier (UUID)"},
                {"name": "npi",                 "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "10-digit National Provider Identifier (unique business key)"},
                {"name": "provider_type",       "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "INDIVIDUAL or ORGANIZATION"},
                {"name": "taxonomy_code",       "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "NUCC Health Care Provider Taxonomy code"},
                {"name": "specialty_description","data_type": "STRING", "nullable": True,  "primary_key": False, "description": "Human-readable specialty (e.g. Internal Medicine)"},
                {"name": "first_name",          "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Provider first name (individual providers only)"},
                {"name": "last_name",           "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Provider last name (individual providers only)"},
                {"name": "organization_name",   "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Organization name (org providers only)"},
                {"name": "credential",          "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Clinical credential: MD, DO, NP, PA, RN, etc."},
                {"name": "address_line1",       "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Practice street address line 1"},
                {"name": "address_city",        "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Practice city"},
                {"name": "address_state",       "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Practice state (2-letter)"},
                {"name": "address_zip",         "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Practice ZIP code"},
                {"name": "phone_number",        "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Primary contact phone number"},
                {"name": "network_status",      "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "IN_NETWORK or OUT_OF_NETWORK"},
                {"name": "contract_effective",  "data_type": "DATE",    "nullable": True,  "primary_key": False, "description": "Date provider contract became effective"},
                {"name": "contract_end",        "data_type": "DATE",    "nullable": True,  "primary_key": False, "description": "Date provider contract ended; NULL if active"},
                {"name": "accepting_new_patients","data_type": "BOOLEAN","nullable": True, "primary_key": False, "description": "Whether provider is accepting new patients"},
                {"name": "quality_score",       "data_type": "DECIMAL(5,2)", "nullable": True, "primary_key": False, "description": "Composite quality score 0.00–100.00"},
                {"name": "created_at",          "data_type": "TIMESTAMP","nullable": False, "primary_key": False, "description": "Row insert timestamp (UTC)"},
                {"name": "updated_at",          "data_type": "TIMESTAMP","nullable": False, "primary_key": False, "description": "Row last-update timestamp (UTC)"},
            ]
        },
        {
            "name": "diagnoses",
            "description": (
                "Bridge table linking claims to ICD-10-CM diagnosis codes. "
                "Each claim may have multiple diagnoses; one is flagged as the principal diagnosis."
            ),
            "estimated_row_count": 24_680_201,
            "columns": [
                {"name": "diagnosis_id",        "data_type": "BIGINT",  "nullable": False, "primary_key": True,  "description": "Surrogate key (auto-increment)"},
                {"name": "claim_id",            "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "FK → claims.claim_id"},
                {"name": "diagnosis_code",      "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "ICD-10-CM diagnosis code (e.g. E11.9 = Type 2 diabetes)"},
                {"name": "diagnosis_description","data_type": "STRING", "nullable": True,  "primary_key": False, "description": "Plain-text description of the ICD-10-CM code"},
                {"name": "code_version",        "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "ICD code version: ICD10, ICD9 (legacy)"},
                {"name": "diagnosis_type",      "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "PRINCIPAL, ADMITTING, or SECONDARY"},
                {"name": "sequence_number",     "data_type": "INT",     "nullable": False, "primary_key": False, "description": "Order of diagnosis on the claim form (1 = principal)"},
                {"name": "poa_indicator",       "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Present on Admission indicator: Y, N, U, W (inpatient only)"},
                {"name": "chronic_flag",        "data_type": "BOOLEAN", "nullable": True,  "primary_key": False, "description": "True if CMS classifies this code as a chronic condition"},
                {"name": "hcc_category",        "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "CMS Hierarchical Condition Category mapping"},
                {"name": "created_at",          "data_type": "TIMESTAMP","nullable": False, "primary_key": False, "description": "Row insert timestamp (UTC)"},
            ]
        },
        {
            "name": "procedures",
            "description": (
                "Bridge table linking claims to CPT/HCPCS procedure codes. "
                "Each row represents one line item on a claim with its associated units and revenue code."
            ),
            "estimated_row_count": 31_102_890,
            "columns": [
                {"name": "procedure_id",        "data_type": "BIGINT",  "nullable": False, "primary_key": True,  "description": "Surrogate key (auto-increment)"},
                {"name": "claim_id",            "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "FK → claims.claim_id"},
                {"name": "line_number",         "data_type": "INT",     "nullable": False, "primary_key": False, "description": "Claim line item number (1-based)"},
                {"name": "procedure_code",      "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "CPT or HCPCS Level II procedure code"},
                {"name": "procedure_description","data_type": "STRING", "nullable": True,  "primary_key": False, "description": "Short description of the procedure"},
                {"name": "code_type",           "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "CPT4, HCPCS, ICD10PCS, or CDT (dental)"},
                {"name": "modifier_1",          "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "First CPT modifier code (e.g. 26=professional component)"},
                {"name": "modifier_2",          "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Second CPT modifier code"},
                {"name": "revenue_code",        "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "UB-04 revenue code (facility claims only)"},
                {"name": "drg_code",            "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "MS-DRG code for inpatient claims"},
                {"name": "service_date",        "data_type": "DATE",    "nullable": False, "primary_key": False, "description": "Date this specific service line was rendered"},
                {"name": "units",               "data_type": "DECIMAL(10,3)", "nullable": False, "primary_key": False, "description": "Quantity of units/services billed on this line"},
                {"name": "unit_type",           "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Unit of measure: UN (units), DA (days), ML, etc."},
                {"name": "billed_amount",       "data_type": "DECIMAL(12,2)", "nullable": False, "primary_key": False, "description": "Amount billed for this procedure line"},
                {"name": "allowed_amount",      "data_type": "DECIMAL(12,2)", "nullable": True,  "primary_key": False, "description": "Payer-allowed amount for this line"},
                {"name": "paid_amount",         "data_type": "DECIMAL(12,2)", "nullable": True,  "primary_key": False, "description": "Amount paid for this line"},
                {"name": "rendering_npi",       "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "NPI of provider who actually rendered the service"},
                {"name": "ndc_code",            "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "National Drug Code (pharmacy/drug claims only)"},
                {"name": "created_at",          "data_type": "TIMESTAMP","nullable": False, "primary_key": False, "description": "Row insert timestamp (UTC)"},
            ]
        },
        {
            "name": "payments",
            "description": (
                "Tracks all payment transactions associated with claims including "
                "insurer payments, patient payments, adjustments, and refunds."
            ),
            "estimated_row_count": 9_874_112,
            "columns": [
                {"name": "payment_id",          "data_type": "STRING",  "nullable": False, "primary_key": True,  "description": "Unique payment transaction identifier (UUID)"},
                {"name": "claim_id",            "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "FK → claims.claim_id"},
                {"name": "payment_type",        "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "INSURER_PAYMENT, PATIENT_PAYMENT, ADJUSTMENT, REFUND, WRITE_OFF"},
                {"name": "payment_amount",      "data_type": "DECIMAL(12,2)", "nullable": False, "primary_key": False, "description": "Payment amount (negative for refunds/adjustments)"},
                {"name": "payment_date",        "data_type": "DATE",    "nullable": False, "primary_key": False, "description": "Date payment was issued or posted"},
                {"name": "payment_method",      "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "EFT, CHECK, CREDIT_CARD, CASH, ERA"},
                {"name": "check_number",        "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Check or EFT trace number"},
                {"name": "remittance_advice_id","data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "835 ERA transaction ID"},
                {"name": "payer_id",            "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "Payer/insurer identifier"},
                {"name": "payer_name",          "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Payer display name"},
                {"name": "adjustment_reason_code","data_type": "STRING","nullable": True,  "primary_key": False, "description": "CARC adjustment reason code (for ADJUSTMENT type)"},
                {"name": "adjustment_group_code","data_type": "STRING", "nullable": True,  "primary_key": False, "description": "ANSI X12 group code: CO, PR, OA, PI, CR"},
                {"name": "is_reconciled",       "data_type": "BOOLEAN", "nullable": False, "primary_key": False, "description": "Whether this payment has been reconciled to the GL"},
                {"name": "gl_posting_date",     "data_type": "DATE",    "nullable": True,  "primary_key": False, "description": "Date payment was posted to the general ledger"},
                {"name": "created_at",          "data_type": "TIMESTAMP","nullable": False, "primary_key": False, "description": "Row insert timestamp (UTC)"},
                {"name": "updated_at",          "data_type": "TIMESTAMP","nullable": False, "primary_key": False, "description": "Row last-update timestamp (UTC)"},
            ]
        },
        {
            "name": "claim_audits",
            "description": (
                "Audit trail capturing every status change and adjudication event "
                "on a claim. Enables full lifecycle tracking and compliance reporting."
            ),
            "estimated_row_count": 42_315_670,
            "columns": [
                {"name": "audit_id",            "data_type": "BIGINT",  "nullable": False, "primary_key": True,  "description": "Surrogate key (auto-increment)"},
                {"name": "claim_id",            "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "FK → claims.claim_id"},
                {"name": "event_type",          "data_type": "STRING",  "nullable": False, "primary_key": False, "description": "Type of event: STATUS_CHANGE, PAYMENT, EDIT, APPEAL, NOTE"},
                {"name": "old_status",          "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Claim status before this event"},
                {"name": "new_status",          "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Claim status after this event"},
                {"name": "event_timestamp",     "data_type": "TIMESTAMP","nullable": False, "primary_key": False, "description": "UTC timestamp of the event"},
                {"name": "actor_id",            "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "User ID or system ID that triggered the event"},
                {"name": "actor_type",          "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "HUMAN, SYSTEM, PAYER_INTERFACE"},
                {"name": "notes",               "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Free-text notes or reason for the event"},
                {"name": "metadata",            "data_type": "MAP<STRING,STRING>", "nullable": True, "primary_key": False, "description": "Arbitrary key-value metadata for the event"},
                {"name": "source_system",       "data_type": "STRING",  "nullable": True,  "primary_key": False, "description": "Originating system: CLAIMS_MGMT, PAYER_PORTAL, BATCH_JOB"},
                {"name": "created_at",          "data_type": "TIMESTAMP","nullable": False, "primary_key": False, "description": "Row insert timestamp (UTC)"},
            ]
        },
    ]
}


# ── Try real Databricks connection ───────────────────────────────────────────
def try_databricks_connection():
    """Return real schema data from Databricks, or None if unavailable."""
    if not all([DATABRICKS_HOST, DATABRICKS_TOKEN, DATABRICKS_HTTP_PATH]):
        print("⚠  Databricks credentials not found — using mock/demo data.")
        return None
    try:
        from databricks.sdk import WorkspaceClient
        print(f"🔌 Connecting to Databricks at {DATABRICKS_HOST} …")
        w = WorkspaceClient(host=DATABRICKS_HOST, token=DATABRICKS_TOKEN)
        tables = list(w.tables.list(catalog_name=CATALOG_NAME, schema_name=SCHEMA_NAME))
        print(f"✅ Connected — found {len(tables)} tables.")
        # Build schema dict from real data
        schema_data = {
            "catalog": CATALOG_NAME,
            "schema":  SCHEMA_NAME,
            "source":  f"Databricks {DATABRICKS_HOST}",
            "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
            "tables": []
        }
        for tbl in tables:
            detail = w.tables.get(f"{CATALOG_NAME}.{SCHEMA_NAME}.{tbl.name}")
            cols = []
            for c in (detail.columns or []):
                cols.append({
                    "name":        c.name,
                    "data_type":   c.type_text or str(c.type_name),
                    "nullable":    c.nullable if c.nullable is not None else True,
                    "primary_key": False,
                    "description": c.comment or "",
                })
            schema_data["tables"].append({
                "name":        tbl.name,
                "description": (detail.comment or ""),
                "estimated_row_count": None,
                "columns": cols,
            })
        return schema_data
    except Exception as exc:
        print(f"⚠  Databricks connection failed ({exc}) — using mock/demo data.")
        return None


# ── Excel generation ─────────────────────────────────────────────────────────
def generate_excel(schema_data: dict, out_path: Path):
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side, GradientFill)
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()

    # Colour palette
    HDR_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
    ALT_FILL   = PatternFill("solid", fgColor="D6E4F0")   # light blue
    PK_FILL    = PatternFill("solid", fgColor="FFF2CC")   # yellow for PK
    SUM_FILL   = PatternFill("solid", fgColor="2E75B6")   # mid blue (summary hdr)
    WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")
    THIN       = Side(border_style="thin", color="B0B0B0")
    BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def hdr_font(size=11, bold=True, color="FFFFFF"):
        return Font(name="Calibri", size=size, bold=bold, color=color)

    def cell_font(size=10, bold=False, color="000000"):
        return Font(name="Calibri", size=size, bold=bold, color=color)

    def wrap_align(h="left", v="center"):
        return Alignment(horizontal=h, vertical=v, wrap_text=True)

    def style_header_row(ws, row, num_cols, fill=HDR_FILL):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill   = fill
            cell.font   = hdr_font()
            cell.border = BORDER
            cell.alignment = wrap_align("center")

    # ── Summary sheet ────────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "📋 Summary"
    ws_sum.sheet_view.showGridLines = False

    # Title block
    ws_sum.merge_cells("A1:F1")
    title_cell = ws_sum["A1"]
    title_cell.value     = f"Healthcare Claims Schema Documentation"
    title_cell.font      = Font(name="Calibri", size=16, bold=True, color="1F4E79")
    title_cell.alignment = wrap_align("center")
    ws_sum.row_dimensions[1].height = 30

    ws_sum.merge_cells("A2:F2")
    sub_cell = ws_sum["A2"]
    sub_cell.value     = (
        f"Catalog: {schema_data['catalog']}  |  Schema: {schema_data['schema']}  |  "
        f"Source: {schema_data['source']}  |  Generated: {schema_data['generated_at']}"
    )
    sub_cell.font      = Font(name="Calibri", size=9, italic=True, color="595959")
    sub_cell.alignment = wrap_align("center")
    ws_sum.row_dimensions[2].height = 18

    ws_sum.row_dimensions[3].height = 6   # spacer

    # Summary header
    sum_headers = ["Table Name", "Row Count (est.)", "Column Count",
                   "Primary Keys", "Has Partitioning", "Description"]
    for col_idx, hdr in enumerate(sum_headers, 1):
        cell = ws_sum.cell(row=4, column=col_idx, value=hdr)
    style_header_row(ws_sum, 4, len(sum_headers), fill=SUM_FILL)
    ws_sum.row_dimensions[4].height = 20

    for row_idx, tbl in enumerate(schema_data["tables"], 5):
        pk_cols    = [c["name"] for c in tbl["columns"] if c.get("primary_key")]
        part_cols  = [c["name"] for c in tbl["columns"] if "partition" in c["name"].lower()]
        row_count  = f"{tbl['estimated_row_count']:,}" if tbl.get("estimated_row_count") else "N/A"
        fill = ALT_FILL if row_idx % 2 == 0 else WHITE_FILL
        vals = [
            tbl["name"],
            row_count,
            len(tbl["columns"]),
            ", ".join(pk_cols) if pk_cols else "—",
            "Yes" if part_cols else "No",
            tbl["description"],
        ]
        for col_idx, val in enumerate(vals, 1):
            cell = ws_sum.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = fill
            cell.font      = cell_font()
            cell.border    = BORDER
            cell.alignment = wrap_align()
        ws_sum.row_dimensions[row_idx].height = 40

    col_widths = [28, 18, 14, 24, 16, 70]
    for i, w in enumerate(col_widths, 1):
        ws_sum.column_dimensions[get_column_letter(i)].width = w

    # ── Per-table sheets ─────────────────────────────────────────────────────
    TABLE_HEADERS = ["#", "Column Name", "Data Type", "Nullable",
                     "Primary Key", "Description / Comment"]
    for tbl in schema_data["tables"]:
        safe_name = tbl["name"][:31]          # Excel tab name limit
        ws = wb.create_sheet(title=f"📄 {safe_name}"[:31])
        ws.sheet_view.showGridLines = False

        # Table title
        ws.merge_cells("A1:F1")
        t = ws["A1"]
        t.value     = f"Table: {schema_data['catalog']}.{schema_data['schema']}.{tbl['name']}"
        t.font      = Font(name="Calibri", size=13, bold=True, color="1F4E79")
        t.alignment = wrap_align("left")
        ws.row_dimensions[1].height = 24

        # Description
        ws.merge_cells("A2:F2")
        d = ws["A2"]
        d.value     = tbl["description"]
        d.font      = Font(name="Calibri", size=9, italic=True, color="595959")
        d.alignment = wrap_align()
        ws.row_dimensions[2].height = 30

        # Stats row
        row_count = f"{tbl['estimated_row_count']:,}" if tbl.get("estimated_row_count") else "N/A"
        ws.merge_cells("A3:F3")
        s = ws["A3"]
        s.value     = f"Estimated rows: {row_count}   |   Columns: {len(tbl['columns'])}"
        s.font      = Font(name="Calibri", size=9, bold=True, color="2E75B6")
        s.alignment = wrap_align("left")
        ws.row_dimensions[3].height = 16

        ws.row_dimensions[4].height = 6   # spacer

        # Column header row
        for col_idx, hdr in enumerate(TABLE_HEADERS, 1):
            cell = ws.cell(row=5, column=col_idx, value=hdr)
        style_header_row(ws, 5, len(TABLE_HEADERS))
        ws.row_dimensions[5].height = 20

        for row_idx, col in enumerate(tbl["columns"], 6):
            is_pk   = col.get("primary_key", False)
            fill    = PK_FILL if is_pk else (ALT_FILL if row_idx % 2 == 0 else WHITE_FILL)
            vals = [
                row_idx - 5,
                col["name"],
                col["data_type"],
                "Yes" if col.get("nullable") else "No",
                "✓ PK" if is_pk else "",
                col.get("description", ""),
            ]
            for col_idx, val in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.fill      = fill
                cell.font      = cell_font(bold=is_pk)
                cell.border    = BORDER
                cell.alignment = wrap_align(
                    "center" if col_idx in (1, 3, 4, 5) else "left"
                )
            ws.row_dimensions[row_idx].height = 18

        # Column widths
        tbl_col_widths = [5, 32, 26, 10, 10, 70]
        for i, w in enumerate(tbl_col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # Freeze panes below header
        ws.freeze_panes = "A6"

    wb.save(out_path)
    print(f"✅ Excel written → {out_path}")


# ── Markdown generation ──────────────────────────────────────────────────────
def generate_markdown(schema_data: dict, out_path: Path):
    lines = []
    now   = schema_data["generated_at"]
    cat   = schema_data["catalog"]
    sch   = schema_data["schema"]
    src   = schema_data["source"]
    tables = schema_data["tables"]

    total_rows = sum(
        t["estimated_row_count"] for t in tables if t.get("estimated_row_count")
    )
    total_cols = sum(len(t["columns"]) for t in tables)

    # ── Header ───────────────────────────────────────────────────────────────
    lines += [
        f"# Healthcare Claims Schema Documentation",
        f"",
        f"> **Catalog:** `{cat}`  **Schema:** `{sch}`",
        f"> ",
        f"> **Source:** {src}",
        f"> **Generated:** {now}",
        f"> **Data Status:** {'⚠️ Sample/Demo Data' if 'sample' in src.lower() else '✅ Live Data'}",
        f"",
        f"---",
        f"",
    ]

    # ── Schema Overview ───────────────────────────────────────────────────────
    lines += [
        f"## Schema Overview",
        f"",
        f"The **`{sch}`** schema is a comprehensive healthcare claims data model that",
        f"supports end-to-end claim lifecycle management — from patient enrollment and",
        f"provider credentialing through claim submission, adjudication, payment, and audit.",
        f"",
        f"| Metric | Value |",
        f"|--------|-------|",
        f"| Tables | {len(tables)} |",
        f"| Total Columns | {total_cols} |",
        f"| Estimated Total Rows | {total_rows:,} |",
        f"",
        f"### Entity-Relationship Summary",
        f"",
        f"```",
        f"patients ──────────────────────────────────┐",
        f"                                            │ patient_id",
        f"providers ─────────────────────────────────┤",
        f"                                            │ provider_id",
        f"                                            ▼",
        f"                                         claims",
        f"                                         /  |  \\",
        f"                              claim_id  /   |   \\ claim_id",
        f"                                       /    |    \\",
        f"                               diagnoses  payments  procedures",
        f"                                                       |",
        f"                                              claim_id |",
        f"                                                       ▼",
        f"                                                 claim_audits",
        f"```",
        f"",
        f"---",
        f"",
    ]

    # ── Table of Contents ────────────────────────────────────────────────────
    lines += ["## Table of Contents", ""]
    for tbl in tables:
        anchor = tbl["name"].lower().replace("_", "-")
        row_count = f"{tbl['estimated_row_count']:,}" if tbl.get("estimated_row_count") else "N/A"
        lines.append(f"- [{tbl['name']}](#{anchor}) — *{row_count} rows*")
    lines += ["", "---", ""]

    # ── Per-table sections ────────────────────────────────────────────────────
    for tbl in tables:
        anchor    = tbl["name"].lower().replace("_", "-")
        pk_cols   = [c["name"] for c in tbl["columns"] if c.get("primary_key")]
        fk_cols   = [c for c in tbl["columns"] if "FK →" in c.get("description", "")]
        part_cols = [c["name"] for c in tbl["columns"] if "partition" in c["name"].lower()]
        nullable_count  = sum(1 for c in tbl["columns"] if c.get("nullable"))
        row_count = f"{tbl['estimated_row_count']:,}" if tbl.get("estimated_row_count") else "N/A"

        lines += [
            f"## `{tbl['name']}`",
            f"",
            f"{tbl['description']}",
            f"",
            f"| Property | Value |",
            f"|----------|-------|",
            f"| Full Name | `{cat}.{sch}.{tbl['name']}` |",
            f"| Estimated Rows | {row_count} |",
            f"| Columns | {len(tbl['columns'])} |",
            f"| Nullable Columns | {nullable_count} |",
            f"| Primary Key(s) | {', '.join(f'`{p}`' for p in pk_cols) if pk_cols else '—'} |",
            f"| Partition Column(s) | {', '.join(f'`{p}`' for p in part_cols) if part_cols else 'None'} |",
            f"",
        ]

        if fk_cols:
            lines += ["**Foreign Keys:**", ""]
            for fk in fk_cols:
                ref = fk["description"].split("FK →")[-1].strip()
                lines.append(f"- `{fk['name']}` → `{ref}`")
            lines.append("")

        # Column table
        lines += [
            "### Columns",
            "",
            "| # | Column Name | Data Type | Nullable | PK | Description |",
            "|---|-------------|-----------|----------|----|-------------|",
        ]
        for i, col in enumerate(tbl["columns"], 1):
            nullable = "✓" if col.get("nullable") else "✗"
            pk       = "🔑" if col.get("primary_key") else ""
            desc     = col.get("description", "").replace("|", "\\|")
            lines.append(
                f"| {i} | `{col['name']}` | `{col['data_type']}` "
                f"| {nullable} | {pk} | {desc} |"
            )

        lines += ["", "---", ""]

    # ── Data Dictionary Appendix ──────────────────────────────────────────────
    lines += [
        "## Appendix: Code Value Dictionaries",
        "",
        "### `claims.claim_status`",
        "",
        "| Value | Description |",
        "|-------|-------------|",
        "| `SUBMITTED` | Claim received by payer, awaiting review |",
        "| `PENDING` | Under review / additional information requested |",
        "| `APPROVED` | Adjudicated and approved for payment |",
        "| `DENIED` | Adjudicated and denied |",
        "| `PAID` | Payment issued to provider |",
        "| `APPEALED` | Provider has filed an appeal |",
        "",
        "### `claims.claim_type`",
        "",
        "| Value | Description |",
        "|-------|-------------|",
        "| `MEDICAL` | Physician or outpatient medical service |",
        "| `DENTAL` | Dental services (CDT codes) |",
        "| `VISION` | Vision/optical services |",
        "| `PHARMACY` | Prescription drug claims |",
        "",
        "### `providers.network_status`",
        "",
        "| Value | Description |",
        "|-------|-------------|",
        "| `IN_NETWORK` | Provider has a contract with the payer |",
        "| `OUT_OF_NETWORK` | No contract; higher cost-sharing for patient |",
        "",
        "### `payments.adjustment_group_code` (ANSI X12)",
        "",
        "| Code | Description |",
        "|------|-------------|",
        "| `CO` | Contractual Obligation — provider write-off |",
        "| `PR` | Patient Responsibility — copay/deductible/coinsurance |",
        "| `OA` | Other Adjustments |",
        "| `PI` | Payer Initiated |",
        "| `CR` | Correction/Reversal |",
        "",
        "---",
        "",
        f"*Documentation generated automatically by the DataEngineer agent · {now}*",
        "",
    ]

    out_path.write_text("\n".join(lines), encoding="utf-8")
    print(f"✅ Markdown written → {out_path}")


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    schema_data = try_databricks_connection() or MOCK_SCHEMA
    schema_data["generated_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")

    xl_path = OUTPUT_DIR / "healthcare_claims_schema.xlsx"
    md_path = OUTPUT_DIR / "healthcare_claims_schema.md"

    print(f"\n📂 Output directory: {OUTPUT_DIR}")
    print(f"📊 Schema: {schema_data['catalog']}.{schema_data['schema']}")
    print(f"📋 Tables: {len(schema_data['tables'])}\n")

    generate_excel(schema_data, xl_path)
    generate_markdown(schema_data, md_path)

    print("\n✅ All files generated successfully.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
